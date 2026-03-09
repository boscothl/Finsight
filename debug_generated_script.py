import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_report():
    """
    Generates an Excel report for system expense claims based on the provided data.
    """
    
    # DATA PAYLOAD (from secure DB pipeline)
    data_payload = """
    [{"id": 1, "merchant": "Uber", "amount_hkd": "500.00", "date": "2026-03-01", "category": "Travel", "status": "approved"}, {"id": 2, "merchant": "Cafe de Coral", "amount_hkd": "200.00", "date": "2026-03-01", "category": "Meal", "status": "approved"}, {"id": 3, "merchant": "Apa Hotel", "amount_hkd": "1000.00", "date": "2026-03-01", "category": "Hotel", "status": "pending"}, {"id": 9, "merchant": "Cathay Pacific", "amount_hkd": "4500.00", "date": "2026-08-10", "category": "Travel Expense", "status": "approved"}, {"id": 10, "merchant": "Starbucks", "amount_hkd": "45.00", "date": "2026-08-12", "category": "Meals & Entertainment", "status": "approved"}, {"id": 11, "merchant": "Adobe Systems", "amount_hkd": "250.00", "date": "2026-08-15", "category": "IT & Software", "status": "approved"}, {"id": 12, "merchant": "IKEA", "amount_hkd": "1200.00", "date": "2026-08-20", "category": "Office Supplies", "status": "approved"}, {"id": 13, "merchant": "Marriott Hotel", "amount_hkd": "3200.00", "date": "2026-08-25", "category": "Travel Expense", "status": "approved"}]
    """
    
    # Parse the data payload
    data = json.loads(data_payload)
    
    # Style Requirements
    style_theme = {"colors": {"primary": "002060", "text": "ffffff"}, "fonts": {"primary": "Arial"}}
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All System Expense Claims"

    # Define headers from the keys of the first data item
    if not data:
        headers = []
    else:
        headers = list(data[0].keys())

    ws.append(headers)

    # Style the header row
    header_font = Font(name=style_theme["fonts"]["primary"], color=style_theme["colors"]["text"], bold=True)
    header_fill = PatternFill(start_color=style_theme["colors"]["primary"], end_color=style_theme["colors"]["primary"], fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Define fills for status
    status_fills = {
        "approved": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "pending": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "rejected": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    }

    # Populate data rows
    for item in data:
        # Convert amount to float for proper formatting
        item['amount_hkd'] = float(item['amount_hkd'])
        row_values = [item.get(h) for h in headers]
        ws.append(row_values)

    # Apply formatting and styles to data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        status_cell = ws.cell(row=row_idx, column=headers.index("status") + 1)
        status = status_cell.value.lower() if status_cell.value else ""
        
        fill_to_apply = status_fills.get(status)

        for cell in row:
            if fill_to_apply:
                cell.fill = fill_to_apply
            
            # Format amount column as currency
            if cell.column == headers.index("amount_hkd") + 1:
                cell.number_format = '"HK$"#,##0.00'

    # Adjust column widths for readability
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Add a buffer to the max length for better spacing
        adjusted_width = (max_length + 4)
        
        # Set a reasonable max width
        if adjusted_width > 50:
            adjusted_width = 50
            
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    output_file_name = "agentic_output_flow.xlsx"
    wb.save(output_file_name)

if __name__ == "__main__":
    generate_report()